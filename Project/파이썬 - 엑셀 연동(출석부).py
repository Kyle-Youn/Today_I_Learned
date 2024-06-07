from openpyxl import load_workbook
from datetime import datetime
import calendar

def get_user_input():
    year = int(input("년도를 입력하세요: "))
    month = int(input("월을 입력하세요: "))
    days = input("수업이 있는 요일을 입력하세요 (예: 월,수,금): ").split(',')
    holiday_days = input("휴일을 입력하세요 (일자만 입력, 예: 5,15, 다중 입력시 쉼표로 구분, 없을 경우 엔터): ").strip()
    if holiday_days:
        holidays = [f"{year}-{month:02d}-{day.strip():02d}" for day in holiday_days.split(',') if day.strip().isdigit()]
    else:
        holidays = []
    return year, month, days, holidays

def process_dates(filepath, year, month, days, holidays):
    # 한국어 요일을 요일 인덱스로 매핑
    day_map = {'월': 0, '화': 1, '수': 2, '목': 3, '금': 4, '토': 5, '일': 6}
    day_numbers = [day_map[day.strip()] for day in days if day.strip() in day_map]

    wb = load_workbook(filepath)
    ws = wb.active

    holidays = set(holidays)  # 휴일을 집합으로 변환

    # 달의 모든 날짜를 생성하고 요일 필터링
    days_in_month = calendar.monthrange(year, month)[1]
    valid_dates = [datetime(year, month, day) for day in range(1, days_in_month + 1)
                   if datetime(year, month, day).weekday() in day_numbers
                   and datetime(year, month, day).strftime('%Y-%m-%d') not in holidays]

    # D2부터 순서대로 날짜를 기입
    for i, valid_date in enumerate(valid_dates[:12]):  # 최대 12개의 날짜만 기입
        cell = ws.cell(row=2, column=4 + i)  # D2는 4번째 열
        cell.value = valid_date.strftime('%Y-%m-%d')

    # 파일 저장
    wb.save(filepath)

# 사용자 입력 받기
year, month, class_days, holidays = get_user_input()
# 파일 경로 설정
file_path = 'C:/Users/hjyou/Desktop/24chul.xlsx'
process_dates(file_path, year, month, class_days, holidays)



# 'C:/Users/hjyou/Desktop/24chul.xlsx'
