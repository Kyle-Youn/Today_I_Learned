from openpyxl import Workbook, load_workbook

# 새 워크북 생성
wb = Workbook()
ws = wb.active

# 데이터 추가
ws['A1'] = "Hello"
ws['B1'] = "World!"

# 엑셀 파일 저장
wb.save("example.xlsx")

# 파일 불러오기
wb2 = load_workbook("example.xlsx")
ws2 = wb2.active

# 특정 셀 값 출력
print(ws2['A1'].value)
print(ws2['B1'].value)
