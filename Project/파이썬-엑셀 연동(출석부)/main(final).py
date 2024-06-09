import tkinter as tk
from tkinter import simpledialog, messagebox
from openpyxl import load_workbook
from datetime import datetime
import calendar
from openpyxl.styles import Font, Border, Side, Alignment

def get_user_input():
    root = tk.Tk()
    root.withdraw()  # GUI 창을 숨깁니다.
    
    year = simpledialog.askinteger("Input", "년도를 입력하세요:", parent=root)
    month = simpledialog.askinteger("Input", "월을 입력하세요:", parent=root)
    days = simpledialog.askstring("Input", "수업이 있는 요일을 입력하세요 (예: 월,수,금):", parent=root)
    holiday_days = simpledialog.askstring("Input", "휴일을 입력하세요 (일자만 입력, 예: 5,15, 다중 입력시 쉼표로 구분, 없을 경우 엔터):", parent=root)

    if holiday_days:
        holidays = [f"{year}-{month:02d}-{int(day.strip()):02d}" for day in holiday_days.split(',') if day.strip().isdigit()]
    else:
        holidays = []
    days = days.split(',') if days else []
    
    root.destroy()
    return year, month, days, holidays

def add_border(cell):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    cell.border = thin_border

def remove_border(cell):
    no_border = Border(left=Side(style=None), right=Side(style=None),
                       top=Side(style=None), bottom=Side(style=None))
    cell.border = no_border

def process_dates(filepath, year, month, days, holidays):
    wb = load_workbook(filepath)
    ws = wb.active
    ws['A1'] = f"{year}학년도 YBL 출석부({month}월)"

    day_map = {'월': 0, '화': 1, '수': 2, '목': 3, '금': 4, '토': 5, '일': 6}
    day_numbers = [day_map[day.strip()] for day in days if day.strip() in day_map]
    holidays = set(holidays)

    days_in_month = calendar.monthrange(year, month)[1]
    valid_dates = [datetime(year, month, day) for day in range(1, days_in_month + 1)
                   if datetime(year, month, day).weekday() in day_numbers
                   and datetime(year, month, day).strftime('%Y-%m-%d') not in holidays]

    font_style = Font(name='맑은 고딕', size=10)
    center_aligned_text = Alignment(horizontal='center')

    # 테두리 제거
    for col in range(4, 16):  # D열부터 O열
        for row in range(2, 24):  # D2부터 O2까지
            remove_border(ws.cell(row=row, column=col))
        for row in range(25, 47):  # D25부터 O25까지
            remove_border(ws.cell(row=row, column=col))

    # 유효한 날짜를 가진 셀 및 아래 21개 셀에 테두리 추가
    for i, valid_date in enumerate(valid_dates[:12]):
        for offset in range(22):  # 아래로 21개 셀까지
            cell = ws.cell(row=2 + offset, column=4 + i)
            cell.value = valid_date.strftime('%m월 %d일') if offset == 0 else None
            cell.font = font_style
            cell.alignment = center_aligned_text
            add_border(cell)

            cell = ws.cell(row=25 + offset, column=4 + i)
            cell.value = valid_date.strftime('%m월 %d일') if offset == 0 else None
            cell.font = font_style
            cell.alignment = center_aligned_text
            add_border(cell)

    wb.save(filepath)
    messagebox.showinfo("완료", "수정이 완료되었습니다!")

year, month, class_days, holidays = get_user_input()
file_path = '24출석부.xlsx'  # 이 경로는 실행 파일이 있는 위치로 설정되었습니다.
process_dates(file_path, year, month, class_days, holidays)